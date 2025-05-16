import os

import openpyxl

base_dir = "docs/mainnet-networks"
wb = openpyxl.load_workbook("Tools.xlsx")
ws = wb["EXP+TOOLS"]

# Получаем список всех папок-проектов, отсортированных по алфавиту
all_projects = [
    name
    for name in sorted(os.listdir(base_dir))
    if os.path.isdir(os.path.join(base_dir, name)) and not name.startswith("template")
]

# Найдём индекс, с которого начинаем (с haqq)
start_from = "haqq"
start_idx = all_projects.index(start_from)
projects_to_process = all_projects[start_idx:]

# Парсим Excel и собираем все строки для каждого проекта
current_project = None
project_rows = {}

for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
    cell_value = row[2].value  # C (project name)
    if cell_value and isinstance(cell_value, str) and cell_value.strip():
        current_project = cell_value.strip().lower().replace(" ", "")
    if current_project:
        if current_project not in project_rows:
            project_rows[current_project] = []
        project_rows[current_project].append(row)

for project in projects_to_process:
    project_title = project.capitalize()
    project_icon = f"{project}-icon.svg"
    project_dir = os.path.join(base_dir, project)
    explorers_path = os.path.join(project_dir, "explorers.mdx")
    tools_path = os.path.join(project_dir, "useful-tools.mdx")

    # Пропуск, если уже есть
    if os.path.exists(explorers_path) or os.path.exists(tools_path):
        print(
            f"[=] Уже есть {project_title}: {explorers_path if os.path.exists(explorers_path) else tools_path}"
        )
        continue

    # Находим ключ в Excel (может быть вариант с дефисами, нижним регистром и т.п.)
    found_key = next((k for k in project_rows if project in k), None)
    if not found_key:
        print(f"[!] Пропущено {project_title}: нет в Excel.")
        continue

    explorers, tools = [], []
    for row in project_rows[found_key]:
        expl_name = row[4].value  # EXP MAINNET (E)
        if expl_name and row[4].hyperlink:
            url = row[4].hyperlink.target
            explorers.append((expl_name, url))
        tool_name = row[5].value  # USEFUL TOOLS (F)
        if tool_name and row[5].hyperlink:
            url = row[5].hyperlink.target
            cat = row[6].value if len(row) > 6 else ""
            desc = row[7].value if len(row) > 7 else ""
            tools.append((tool_name, url, cat, desc))

    explorers_header = f"""---
title: 'Explorers List'
sidebar_position: 10
hide_title: true
hide_table_of_contents: false
---

import PageTitle from '@site/src/components/PageTitle';

<PageTitle iconUrl="img/{project_icon}" title="{project_title} Explorers List" />

## A list of blockchain explorers designed for {project_title}

| Name             | Link |
| ---------------- | ---- |
"""
    useful_tools_header = f"""---
title: 'Useful Tools'
sidebar_position: 11
hide_title: true
hide_table_of_contents: false
---

import PageTitle from '@site/src/components/PageTitle';

<PageTitle iconUrl="img/{project_icon}" title="{project_title} Useful Tools" />

## This is a list of useful tools designed for {project_title}

| Name | Link | Category | Description |
|------|------|----------|-------------|
"""

    # Генерируем ссылки как ты просил
    explorers_table = "\n".join([f"| {n} | [{u}]({u}) |" for n, u in explorers if u])
    useful_tools_table = "\n".join(
        [f"| {n} | [{u}]({u}) | {c or ''} | {d or ''} |" for n, u, c, d in tools if u]
    )

    with open(explorers_path, "w") as f:
        f.write(explorers_header + explorers_table + "\n")

    with open(tools_path, "w") as f:
        f.write(useful_tools_header + useful_tools_table + "\n")

    print(f"[+] Создано для {project_title}!")

print("Готово!")
